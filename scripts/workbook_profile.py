"""Краткая воспроизводимая проверка структуры и кэшированных ошибок Excel-книги.

Запуск:
  python scripts/workbook_profile.py "..\\Актуальная версия — моделирование 09.06.26.xlsx"
"""

from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def main(source: str) -> None:
    source_path = Path(source)
    if not source_path.exists():
        raise SystemExit(f"Файл не найден: {source_path}")

    formulas = load_workbook(source_path, read_only=False, data_only=False)
    cached = load_workbook(source_path, read_only=False, data_only=True)
    print(f"Источник: {source_path.name}")
    print(f"Листов: {len(formulas.sheetnames)}")

    total_errors = Counter()
    for name in formulas.sheetnames:
        formula_sheet = formulas[name]
        cached_sheet = cached[name]
        formula_count = 0
        errors = Counter()
        for row in formula_sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1
                    cached_value = cached_sheet[cell.coordinate].value
                    if isinstance(cached_value, str) and cached_value.startswith("#"):
                        errors[cached_value] += 1
                        total_errors[cached_value] += 1
        print(
            f"- {name}: {formula_sheet.max_row}x{formula_sheet.max_column}, "
            f"формул {formula_count}, ошибок {sum(errors.values())} {dict(errors)}"
        )

    print(f"Всего кэшированных ошибок: {sum(total_errors.values())} {dict(total_errors)}")
    print("Важно: openpyxl не пересчитывает формулы; после исправления откройте книгу в Excel и повторите проверку.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Передайте путь к .xlsx-файлу")
    main(sys.argv[1])
