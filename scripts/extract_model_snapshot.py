"""Create a compact, auditable JSON snapshot from the supplied Excel workbooks.

Run this script from the folder containing both source workbooks.  The workbook
with the larger number of sheets is treated as the actual-model workbook; the
other workbook supplies the page/tab structure.
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> float:
    if isinstance(value, bool):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


def period(value: Any) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    return None


def load_workbooks(source_dir: Path):
    paths = sorted(source_dir.glob("*.xlsx"))
    if len(paths) < 2:
        raise FileNotFoundError("В исходной папке должны находиться две книги .xlsx.")
    books = [(path, openpyxl.load_workbook(path, data_only=True, read_only=False)) for path in paths]
    books.sort(key=lambda item: len(item[1].worksheets))
    return books[0], books[-1]


def dates_in_row(sheet, row: int) -> dict[int, str]:
    return {column: item for column in range(1, sheet.max_column + 1) if (item := period(sheet.cell(row, column).value))}


def extract_finance(summary):
    year_columns = {2024: 8, 2025: 16, 2026: 21}
    source_rows = [6, 7, 8, 9, 10, 11, 12, 15]
    lines = []
    by_label = {}
    for row in source_rows:
        label = text(summary.cell(row, 2).value)
        values = {str(year): number(summary.cell(row, column).value) for year, column in year_columns.items()}
        line = {"label": label, "values": values}
        lines.append(line)
        by_label[label] = values

    work = by_label.get("Объем работ в т.ч. НДС", {})
    vat_income = by_label.get("НДС", {})
    cost = by_label.get("Себестоимость работ всего", {})
    vat_cost = by_label.get("НДС всего", {})
    operating = {
        str(year): (work.get(str(year), 0) - vat_income.get(str(year), 0)) - (cost.get(str(year), 0) - vat_cost.get(str(year), 0))
        for year in year_columns
    }
    lines.append({"label": "Операционная разница без НДС", "values": operating, "derived": True})
    return {"years": [2024, 2025, 2026], "lines": lines, "operating": operating}


def extract_cash_receipts(sheet):
    header_dates = dates_in_row(sheet, 2)
    records = []
    project = ""
    contract_period = ""
    record_number = 1
    for row in range(3, sheet.max_row + 1):
        project_cell = text(sheet.cell(row, 1).value)
        if project_cell:
            project = project_cell
        period_cell = text(sheet.cell(row, 2).value)
        if period_cell:
            contract_period = period_cell
        if text(sheet.cell(row, 4).value).lower() != "всего" or not project:
            continue
        monthly = [
            {"period": item, "amount": number(sheet.cell(row, column).value)}
            for column, item in header_dates.items()
            if number(sheet.cell(row, column).value)
        ]
        total = sum(item["amount"] for item in monthly)
        if not total:
            continue
        records.append(
            {
                "id": record_number,
                "project": project,
                "contractPeriod": contract_period,
                "total": total,
                "monthly": monthly,
            }
        )
        record_number += 1
    return records


def extract_subcontracts(sheet):
    header_dates = dates_in_row(sheet, 13)
    records = []
    current_project = ""
    record_number = 1
    for row in range(1, sheet.max_row + 1):
        marker = sheet.cell(row, 1).value
        marker_text = text(marker)
        if marker_text and ":" in marker_text and "№" not in marker_text:
            current_project = marker_text
        vendor = text(sheet.cell(row, 2).value)
        if not isinstance(marker, (int, float)) or not vendor:
            continue
        monthly = [
            {"period": item, "amount": number(sheet.cell(row, column).value)}
            for column, item in header_dates.items()
            if number(sheet.cell(row, column).value)
        ]
        monthly_amount = sum(item["amount"] for item in monthly)
        rate = number(sheet.cell(row, 4).value)
        if not monthly_amount:
            continue
        records.append(
            {
                "id": record_number,
                "project": current_project or "Не указан",
                "vendor": vendor,
                "subject": text(sheet.cell(row, 3).value),
                "resource": text(sheet.cell(row, 4).value),
                "rate": rate,
                "amount": monthly_amount,
                "estimatedHours": round(monthly_amount / rate, 1) if rate else None,
                "monthly": monthly,
            }
        )
        record_number += 1
    return records


def extract_staff_resources(sheet):
    records = []
    resource_group = "Внутренние ресурсы"
    record_number = 1
    for row in range(7, sheet.max_row + 1):
        marker = text(sheet.cell(row, 1).value)
        if marker and not isinstance(sheet.cell(row, 1).value, (int, float)) and not text(sheet.cell(row, 2).value):
            resource_group = marker
        employee = text(sheet.cell(row, 2).value)
        project_name = text(sheet.cell(row, 3).value)
        role = text(sheet.cell(row, 4).value)
        hours = number(sheet.cell(row, 5).value)
        cost = number(sheet.cell(row, 18).value)
        if not employee or not project_name or not role or (not hours and not cost):
            continue
        records.append(
            {
                "id": record_number,
                "group": resource_group,
                "employee": employee,
                "project": project_name,
                "role": role,
                "hours": hours,
                "cost": cost,
            }
        )
        record_number += 1
    return records


def extract_team(staff_sheet):
    role_rows = []
    for row in range(6, 15):
        role = text(staff_sheet.cell(row, 1).value)
        if role:
            role_rows.append(
                {
                    "role": role,
                    "september": number(staff_sheet.cell(row, 2).value),
                    "october": number(staff_sheet.cell(row, 3).value),
                    "november": number(staff_sheet.cell(row, 4).value),
                    "december": number(staff_sheet.cell(row, 5).value),
                }
            )
    roster = []
    for row in range(20, staff_sheet.max_row + 1):
        employee = text(staff_sheet.cell(row, 2).value)
        if employee:
            roster.append(
                {
                    "employee": employee,
                    "project": text(staff_sheet.cell(row, 3).value),
                    "role": text(staff_sheet.cell(row, 4).value),
                }
            )
    return {"roles": role_rows, "roster": roster}


def extract_reference(reference_sheet):
    roles = [text(reference_sheet.cell(row, 2).value) for row in range(3, 17)]
    projects = [text(reference_sheet.cell(row, 2).value) for row in range(19, 29)]
    return {"roles": [item for item in roles if item], "projects": [item for item in projects if item]}


def build_snapshot(source_dir: Path) -> dict[str, Any]:
    (structure_path, structure_book), (model_path, model_book) = load_workbooks(source_dir)
    summary, cash, plan, resources, staff, reference = (
        model_book.worksheets[0],
        model_book.worksheets[2],
        model_book.worksheets[4],
        model_book.worksheets[5],
        model_book.worksheets[11],
        model_book.worksheets[13],
    )
    staff_records = extract_staff_resources(resources)
    projects = sorted({item["project"] for item in extract_cash_receipts(cash)} | {item["project"] for item in staff_records})
    return {
        "source": {
            "structureWorkbook": structure_path.name,
            "dataWorkbook": model_path.name,
            "asOf": "2026-06-09",
            "note": "Значения выгружены из исходной модели; расчётные поля в интерфейсе помечаются отдельно.",
        },
        "tabs": structure_book.sheetnames,
        "projects": projects,
        "finance": extract_finance(summary),
        "cashReceipts": extract_cash_receipts(cash),
        "subcontracts": extract_subcontracts(plan),
        "staffResources": staff_records,
        "team": extract_team(staff),
        "reference": extract_reference(reference),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()
    snapshot = build_snapshot(Path.cwd())
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Created {args.out}")


if __name__ == "__main__":
    main()
