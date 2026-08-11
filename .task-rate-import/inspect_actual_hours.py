import json
import sys

import openpyxl


source_path = sys.argv[1]
workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
result = []
for sheet in workbook.worksheets:
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(max_row=min(sheet.max_row, 24), values_only=True), start=1):
        values = list(row)
        if any(value is not None and str(value).strip() for value in values):
            rows.append({"row": row_number, "values": values})
    result.append({"title": sheet.title, "maxRow": sheet.max_row, "maxColumn": sheet.max_column, "sample": rows})
print(json.dumps(result, ensure_ascii=False, default=str))
