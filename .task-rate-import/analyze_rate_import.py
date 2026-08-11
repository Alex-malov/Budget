import json
import re
import sys
from urllib.request import urlopen

import openpyxl


def normalize(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def number(value):
    if value is None or str(value).strip() == "":
        return 0.0
    return float(value)


source_path = sys.argv[1]
api_url = sys.argv[2]
workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
sheet = workbook["Ставки 26"]

source = {}
duplicate_conflicts = []
duplicate_rows = []
for row_number, row in enumerate(sheet.iter_rows(min_row=5, max_row=129, values_only=True), start=5):
    name = str(row[1] or "").strip()
    if not name:
        continue
    rate = [number(value) for value in row[2:14]]
    attraction = [number(value) for value in row[15:27]]
    record = {"name": name, "rate": rate, "attraction": attraction, "row": row_number}
    key = normalize(name)
    if key in source:
        duplicate_rows.append({"name": name, "rows": [source[key]["row"], row_number]})
        if source[key]["rate"] != rate or source[key]["attraction"] != attraction:
            duplicate_conflicts.append({"name": name, "rows": [source[key]["row"], row_number]})
        continue
    source[key] = record

with urlopen(api_url + "/api/references") as response:
    directories = json.load(response)["directories"]

resource_index = {}
for record in directories["resources"]["records"]:
    for value in [record.get("name", "")] + record.get("sourceValues", []):
        key = normalize(value)
        if key:
            resource_index.setdefault(key, []).append(record)

matched = []
missing = []
ambiguous = []
for key, source_record in source.items():
    candidates = resource_index.get(key, [])
    unique = {candidate["id"]: candidate for candidate in candidates}
    if not unique:
        missing.append(source_record["name"])
    elif len(unique) > 1:
        ambiguous.append({"name": source_record["name"], "resourceIds": sorted(unique)})
    else:
        resource = next(iter(unique.values()))
        matched.append({"source": source_record["name"], "resourceId": resource["id"], "resource": resource["name"], "vendor": resource.get("vendor", ""), "providerType": resource.get("providerType", "")})

print(json.dumps({
    "sourceCount": len(source),
    "sourceRows": len([row for row in sheet.iter_rows(min_row=5, max_row=129, values_only=True) if row[1] is not None and str(row[1]).strip()]),
    "duplicates": duplicate_rows,
    "duplicateConflicts": duplicate_conflicts,
    "matchedCount": len(matched),
    "missing": missing,
    "ambiguous": ambiguous,
    "matched": matched,
    "currentVendorLt": [record for record in directories["vendors"]["records"] if normalize(record.get("name")) == "лт"],
    "currentProviderStaff": [record for record in directories["providers"]["records"] if normalize(record.get("name")) == "штат"]
}, ensure_ascii=False))
