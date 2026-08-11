import json
import sys

import openpyxl


source_path = sys.argv[1]
workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
summary = []
for sheet in workbook.worksheets:
    summary.append({"title": sheet.title, "max_row": sheet.max_row, "max_column": sheet.max_column})

target = workbook["Ставки 26"]
rows = []
for row_number, row in enumerate(target.iter_rows(values_only=True), start=1):
    values = list(row)
    if any(value is not None and str(value).strip() for value in values):
        rows.append({"row": row_number, "values": values})

print(json.dumps({"sheets": summary, "rates_rows": rows}, ensure_ascii=False, default=str))
